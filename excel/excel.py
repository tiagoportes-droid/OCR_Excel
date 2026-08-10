"""
Leitura e escrita no Excel com openpyxl (thread-safe).
Nunca substitui a planilha inteira. Preserva formulas, estilos, filtros.
"""

from __future__ import annotations

import threading
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from config import EXCEL_PATH
from excel.mapping import CAMPOS_OBRIGATORIOS, mapear_colunas
from logger import log

# Lock global para serializar escrita no Excel
_excel_lock = threading.Lock()


def _encontrar_cabecalho(ws: Worksheet) -> tuple[int, list[str]]:
    """
    Localiza a linha de cabecalho da planilha.
    Percorre as primeiras 10 linhas procurando uma com pelo menos 3 celulas preenchidas.
    Retorna (numero_da_linha, lista_de_valores_do_cabecalho).
    """
    for row_num in range(1, 11):
        valores: list[str] = []
        for cell in ws[row_num]:
            valores.append(str(cell.value).strip() if cell.value is not None else "")
        nao_vazios = sum(1 for v in valores if v)
        if nao_vazios >= 3:
            return (row_num, valores)

    # Fallback: usar primeira linha
    valores = []
    for cell in ws[1]:
        valores.append(str(cell.value).strip() if cell.value is not None else "")
    return (1, valores)


def _encontrar_ultima_linha(ws: Worksheet, header_row: int) -> int:
    """Encontra a ultima linha preenchida abaixo do cabecalho."""
    max_row = ws.max_row or header_row
    for row_num in range(max_row, header_row, -1):
        for cell in ws[row_num]:
            if cell.value is not None:
                return row_num
    return header_row


def _garantir_planilha_existe(caminho: Path) -> None:
    """Cria a planilha Excel se nao existir, com cabecalhos padrao."""
    if caminho.exists():
        return
    log.info(f"Planilha nao encontrada. Criando: {caminho}")
    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Dados")
    else:
        ws.title = "Dados"
    cabecalhos = [
        "Data", "Hora", "Tipo", "Cliente", "Local", "Serviço",
        "Valor", "Pagador", "Recebedor", "Empresa", "Nome", "CPF", "CNPJ",
        "Banco", "Agência", "Conta", "PIX", "Código", "Autenticação",
        "Endereço", "Cidade", "Estado", "CEP", "Telefone", "E-mail",
        "Descrição", "Direção", "Confiança", "Observações",
    ]
    for col_idx, cab in enumerate(cabecalhos, start=1):
        ws.cell(row=1, column=col_idx, value=cab)
    wb.save(str(caminho))
    log.info("Planilha criada com cabecalhos padrao.")


def escrever_no_excel(
    dados: dict,
    caminho: Path | None = None,
    aba: str | None = None,
) -> bool:
    """
    Escreve uma nova linha na planilha Excel.

    Thread-safe: utiliza lock para garantir acesso exclusivo.
    Preserva formulas, estilos, filtros, tabelas, validacoes.
    Retorna True se a escrita foi bem-sucedida.
    """
    caminho = caminho or EXCEL_PATH

    with _excel_lock:
        try:
            _garantir_planilha_existe(caminho)

            wb = load_workbook(str(caminho))

            # Selecionar aba
            if aba and aba in wb.sheetnames:
                ws = wb[aba]
            else:
                ws = wb.active

            if ws is None:
                log.error("Planilha sem aba ativa.")
                wb.close()
                return False

            # Encontrar cabecalho
            header_row, cabecalhos = _encontrar_cabecalho(ws)

            # Mapear colunas
            mapa = mapear_colunas(cabecalhos)

            # Verificar campos obrigatorios
            faltando = [c for c in CAMPOS_OBRIGATORIOS if mapa.get(c) is None]
            if faltando:
                log.error(f"Colunas obrigatorias nao encontradas: {faltando}. Abortando escrita.")
                wb.close()
                return False

            # Encontrar proxima linha
            ultima_linha = _encontrar_ultima_linha(ws, header_row)
            nova_linha = ultima_linha + 1

            # Escrever dados
            campos_escritos = 0
            for campo, col_idx in mapa.items():
                if col_idx is None:
                    continue
                valor = dados.get(campo)
                if valor is not None:
                    ws.cell(row=nova_linha, column=col_idx + 1, value=valor)
                    campos_escritos += 1

            if campos_escritos == 0:
                log.warning("Nenhum campo escrito no Excel - dados vazios.")
                wb.close()
                return False

            wb.save(str(caminho))
            wb.close()

            log.info(f"Excel atualizado: linha {nova_linha}, {campos_escritos} campos escritos em {caminho.name}")
            return True

        except PermissionError:
            log.error(f"Excel esta aberto ou sem permissao de escrita: {caminho}. Feche o arquivo e tente novamente.")
            return False
        except Exception as exc:
            log.error(f"Erro ao escrever no Excel {caminho}: {exc}", exc_info=True)
            return False