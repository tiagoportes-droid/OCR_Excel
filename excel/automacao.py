"""
Automacao auxiliar do Excel: verificacoes, listagem de abas, contagem de linhas.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from config import EXCEL_PATH
from logger import log


def verificar_excel_acessivel(caminho: Path | None = None) -> bool:
    """Verifica se o arquivo Excel pode ser aberto para escrita."""
    caminho = caminho or EXCEL_PATH
    if not caminho.exists():
        return True
    try:
        wb = load_workbook(str(caminho))
        wb.close()
        return True
    except PermissionError:
        log.warning(f"Excel nao acessivel (possivelmente aberto): {caminho}")
        return False
    except Exception as exc:
        log.error(f"Erro ao verificar Excel: {exc}")
        return False


def obter_abas(caminho: Path | None = None) -> list[str]:
    """Retorna lista de nomes das abas da planilha."""
    caminho = caminho or EXCEL_PATH
    if not caminho.exists():
        return []
    try:
        wb = load_workbook(str(caminho), read_only=True)
        nomes = wb.sheetnames
        wb.close()
        return nomes
    except Exception as exc:
        log.error(f"Erro ao listar abas: {exc}")
        return []


def contar_linhas(caminho: Path | None = None, aba: str | None = None) -> int:
    """Retorna o numero de linhas preenchidas na aba."""
    caminho = caminho or EXCEL_PATH
    if not caminho.exists():
        return 0
    try:
        wb = load_workbook(str(caminho), read_only=True)
        if aba and aba in wb.sheetnames:
            ws = wb[aba]
        else:
            ws = wb.active
        max_row = ws.max_row or 0 if ws else 0
        wb.close()
        return max_row
    except Exception as exc:
        log.error(f"Erro ao contar linhas: {exc}")
        return 0