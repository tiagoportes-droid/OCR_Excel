"""Testes para leitura/escrita do Excel."""

import sys
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from excel.mapping import mapear_colunas, ALIASES
from excel.excel import escrever_no_excel, _encontrar_cabecalho, _encontrar_ultima_linha


class TestMapeamentoColunas:
    def test_mapeamento_basico(self):
        cabecalhos = ["Data", "Hora", "Valor", "Cliente", "Serviço"]
        mapa = mapear_colunas(cabecalhos)
        assert mapa["data"] == 0
        assert mapa["hora"] == 1
        assert mapa["valor"] == 2
        assert mapa["cliente"] == 3

    def test_mapeamento_alias(self):
        cabecalhos = ["Nome Cliente", "Valor Pago"]
        mapa = mapear_colunas(cabecalhos)
        assert mapa["cliente"] == 0
        assert mapa["valor"] == 1

    def test_coluna_nao_encontrada(self):
        cabecalhos = ["Data", "Hora"]
        mapa = mapear_colunas(cabecalhos)
        assert mapa.get("valor") is None

    def test_cabecalho_vazio(self):
        cabecalhos = ["", "", ""]
        mapa = mapear_colunas(cabecalhos)
        for campo in mapa:
            assert mapa[campo] is None


class TestEscreverExcel:
    def test_escrita_basica(self, tmp_path):
        caminho = tmp_path / "teste.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Data")  # type: ignore
        ws.cell(row=1, column=2, value="Valor")  # type: ignore
        ws.cell(row=1, column=3, value="Cliente")  # type: ignore
        wb.save(str(caminho))
        wb.close()

        dados = {"data": "15/03/2024", "valor": 100.50, "cliente": "Joao"}
        resultado = escrever_no_excel(dados, caminho=caminho)
        assert resultado is True

        wb = load_workbook(str(caminho))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "15/03/2024"  # type: ignore
        assert ws.cell(row=2, column=2).value == 100.50  # type: ignore
        assert ws.cell(row=2, column=3).value == "Joao"  # type: ignore
        wb.close()

    def test_preserva_dados_existentes(self, tmp_path):
        caminho = tmp_path / "teste.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Valor")  # type: ignore
        ws.cell(row=1, column=2, value="Cliente")  # type: ignore
        ws.cell(row=2, column=1, value=50.0)  # type: ignore
        ws.cell(row=2, column=2, value="Maria")  # type: ignore
        wb.save(str(caminho))
        wb.close()

        dados = {"valor": 200.0, "cliente": "Pedro"}
        escrever_no_excel(dados, caminho=caminho)

        wb = load_workbook(str(caminho))
        ws = wb.active
        assert ws.cell(row=2, column=1).value == 50.0  # type: ignore
        assert ws.cell(row=2, column=2).value == "Maria"  # type: ignore
        assert ws.cell(row=3, column=1).value == 200.0  # type: ignore
        assert ws.cell(row=3, column=2).value == "Pedro"  # type: ignore
        wb.close()

    def test_criacao_automatica(self, tmp_path):
        caminho = tmp_path / "novo.xlsx"
        assert not caminho.exists()

        dados = {"valor": 100.0, "data": "01/01/2024"}
        resultado = escrever_no_excel(dados, caminho=caminho)
        assert resultado is True
        assert caminho.exists()